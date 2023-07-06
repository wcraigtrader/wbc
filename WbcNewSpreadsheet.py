# ----- Copyright (c) 2010-2022 by W. Craig Trader ---------------------------------
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published
# by the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

import logging
import sys
from collections import OrderedDict
from datetime import datetime, time, timedelta

import openpyxl

from WbcCalendars import WbcWebcal
from WbcSpreadsheet import WbcRow, WbcSchedule
from WbcUtility import parse_value, sheet_value, text_to_datetime, round_up_timedelta

if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG)
    logging.getLogger('requests').setLevel(logging.WARN)

LOG = logging.getLogger('WbcNewSpreadsheet')

MIDNIGHT = time(0, 0, 0)




class WbcNewRow(WbcRow):
    """
    New format spreadsheet:

    Date
    """
    KEYS = ['Date', 'Day', 'Day Code', 'Time', 'Event Code', 'Event', 'Type', 'Round/Heat', 'Prize Level', 'Class',
            'Format', 'Style', 'Duration', 'Location', 'GM', 'Category']
    FIELDS = ['Date', 'Day', 'DayCode', 'Time', 'Code', 'Name', 'Type', 'RType', 'Prize', 'Class', 'Format', 'Style',
              'Duration', 'Location', 'GM', 'Category']

    def __init__(self, *args):
        self.keymap = OrderedDict(list(zip(self.KEYS, self.FIELDS)))
        self.FIELDS.append('Continuous')
        WbcRow.__init__(self, *args)

    def readrow(self, *args):
        """Custom implementation of readrow to handle XLS-formatted rows"""

        labels = args[0]
        row = args[1]
        datemode = args[2]

        for i in range(len(labels)):
            label = labels[i]
            if label in self.keymap:
                key = self.keymap[label]
                try:
                    val = parse_value(row[i])
                    self.__setattr__(key, val)
                except ValueError as e:
                    raise ValueError(e.message + ' for ' + labels[i])

        self.event = self.name

    def initialize(self):
        if self.line in [25]:
            pass

        # Excel library 'helps' us by treating numeric strings as numbers, not text
        if isinstance(self.code, float):
            self.code = str(int(self.code))
            
        if self.code and self.code in self.meta.MISCODES:
            self.code = self.meta.MISCODES[self.code]

        # Capture metadata that used to be hand-jammed
        if self.code not in self.meta.names:
            self.meta.names[self.code] = self.event
            self.meta.codes[self.event] = self.code
            self.meta.tourneys.append(self.code)

        # Pseudo Code for non-tournament events
        if not self.code:
            self.code = self.type
        if self.type in ['Meeting', 'Services']:
            self.code = 'Seminar'

        # Create old-style fields
        self.continuous = 'Y' if self.style == 'Continuous' else ''
        if self.rtype:
            self.name += ' ' + self.rtype
        # self.rtype = self.round if self.round else None

        # Clean up dates and times, because people still refuse to use date and time fields in Excel
        if isinstance(self.date, datetime):
            self.datetime = self.date
        if isinstance(self.date, str):
            self.datetime = text_to_datetime(self.date)

        if isinstance(self.time, str) or isinstance(self.time, float):
            t = float(self.time) * 60
            h = int(t / 60)
            m = int(t % 60)
            if h > 23:
                h = h - 24
                self.datetime = self.datetime + timedelta(days=1)
            self.datetime = self.datetime.replace(hour=h, minute=m)
            self.time = time(h, m)

        if self.duration:
            try:
                l = timedelta(minutes=60 * float(self.duration))
                self.length = round_up_timedelta(l)
            except:
                LOG.error("Invalid duration (%s) on %s", self.duration, self)
                self.length = timedelta(minutes=0)
        else:
            self.length = timedelta(minutes=0)

        self.date = self.datetime.replace(hour=0, minute=0, second=0, microsecond=0)
        offset = (self.date - self.meta.first_day).days

        if not hasattr(self, 'day'):
            self.day = self.meta.day_names[offset]

        if self.day != self.meta.day_names[offset] and self.time != MIDNIGHT:
            LOG.error("Mismatched day name (%s) on %s", self.day, self)

        if self.daycode != self.meta.day_codes[offset] and self.time != MIDNIGHT:
            LOG.error("Mismatched day code (%s) on %s", self.daycode, self)

        if self.line in [220, 479, 494, 502, 741, 1084, 1103]:
            pass


class WbcNewSchedule(WbcSchedule):
    """
    Starting in 2018, the schedule format was revised so that for each event,
    there was one line for each scheduled session.
    """

    TABS = [
        'by Game',
        'Chronological Website',
        'Chronlogical Website',
    ]

    SPECIALS = {
        'Auction': 'Auction Store',
        'Demos': 'Demonstrations',
        'Juniors': 'Junior Events',
        'Open Gaming': 'Open Gaming',
        'Registration': 'Registration',
        'Seminar': 'Meetings and Seminars',
        'Shuttle': 'Airport Shuttle',
        'Vendors': 'Vendors',
    }

    def __init__(self, *args):
        WbcSchedule.__init__(self, *args)

    def load_events(self):
        """
        Read a new-format Excel spreadsheet and generate WBC events for each row.
        """

        LOG.debug('Initializing metadata for special categories of events')
        for code, name in self.SPECIALS.items():
            self.meta.names[code] = name
            self.meta.codes[name] = code
            self.meta.special = list(self.SPECIALS.keys())

        LOG.debug('Reading new-format Excel spreadsheet from %s', self.filename)

        # Load the Excel workbook
        book = openpyxl.load_workbook(self.filename)

        # Locate the data sheet in the available worksheets
        for tab in self.TABS:
            if tab in book.sheetnames:
                sheet = book[tab]
                break
        else:
            sheet = book.active
            LOG.warning('Unable to locate data sheet by name, using %s', sheet.title)

        # Locate header row (first column named 'Date')
        rbase = 1
        cbase = 1
        nrows = sheet.max_row
        ncols = sheet.max_column

        header_row = 1
        key = sheet_value(sheet, header_row, cbase)
        while key != 'Date' and header_row < nrows:
            header_row += 1
            key = sheet_value(sheet, header_row, cbase)

        if header_row >= nrows:
            raise ValueError('Did not find header row in %s' % self.filename)

        # Read header names
        header = []
        for header_col in range(cbase, ncols+cbase):
            key = None
            try:
                key = sheet_value(sheet, header_row, header_col)
                if key:
                    header.append(key)
                else:
                    raise ValueError('Invalid Column Header %d (%s)' % (header_col, key))
            except Exception:
                raise ValueError('Unable to parse Column Header %d (%s)' % (header_col, key))

        # Scan Date column looking for earliest date (should be First Friday)
        for data_row in range(header_row + 1, nrows+rbase):
            row_date = text_to_datetime(sheet_value(sheet, data_row, cbase))
            self.meta.check_date(row_date)

        # Read data rows
        rows = list(sheet.rows)
        for data_row in range(header_row + 1, nrows+rbase):
            # if self.meta.verbose:
            #     LOG.debug('Reading row %d', data_row + 1)

            try:
                sheet_row = rows[data_row-rbase]
                event_row = WbcNewRow(self, data_row, header, sheet_row, None)

                code = event_row.code
                if not code:
                    LOG.warning("Skipping spreadsheet row %d: no code: %s", data_row, event_row )
                    continue
                elif code in self.events:
                    self.events[code].append(event_row)
                else:
                    self.events[code] = [event_row]

                calendar = self.calendars.get_or_create_event_calendar(code)
                self.calendars.create_event(calendar, event_row, replace=False)

            except Exception as e:
                LOG.exception("Skipping spreadsheet row %d:", data_row)
                # exc_type, exc_obj, exc_tb = sys.exc_info()
                # LOG.error("On line %d, skipping spreadsheet row %d: %s(%s)", exc_tb.tb_lineno, data_row, e.__class__.__name__, e)
                pass

    # ----- Testing --------------------------------------------------------------


if __name__ == '__main__':
    from WbcMetadata import WbcMetadata

    meta = WbcMetadata()
    wbc_calendars = WbcWebcal(meta)

    wbc_schedule = WbcNewSchedule(meta, wbc_calendars)
    wbc_schedule.report_unprocessed_events()

    wbc_calendars.create_all_calendars()

    wbc_schedule.write_all_files()
